import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
from datetime import datetime

data_atual = datetime.now().strftime("%Y%m%d")

ficheiro = r"C:\Users\Eurico\Desktop\EDIT\Ambiente\Projeto1\sales.parquet" 
df = pd.read_parquet(ficheiro)

#corrigir erro do datetime
df["PurchaseDate"] = pd.to_datetime(df["PurchaseDate"], errors="coerce")

#kpis
receita_total = df["TotalPrice"].sum()
media_vendas = df["TotalPrice"].mean()
total_transacoes = df["TransactionID"].nunique()
total_itens_vendidos = df["Quantity"].sum()
vendas_por_categoria = df.groupby("Category")["TotalPrice"].sum()


#top_produtos = df.groupby(["Region", "ProductName"])["ProductName"].count().sort_values(ascending=False).head(20)
top_produtos = df.groupby(["Region", "ProductName"])["Quantity"].sum().reset_index()
top_produtos = top_produtos.sort_values(by=["Region", "Quantity"], ascending=[True, False])
top_produtos = top_produtos.groupby("Region").head(5).reset_index(drop=True)

vendas_por_regiao = df.groupby("Region")["TotalPrice"].sum()
media_quantidade_vendida = df["Quantity"].mean()
vendas_por_metodo_pagamento = df.groupby("PaymentMethod")["TotalPrice"].sum()
invoice_medio_por_cliente = df.groupby("CustomerID")["TotalPrice"].sum().mean()
vendas_acima_5000 = df[df["TotalPrice"] > 5000].shape[0]
vendas_por_categoria_regiao = df.groupby(["Category", "Region"])["TotalPrice"].sum()
transacoes_por_periodo = df.groupby(df["PurchaseDate"].dt.to_period("D")).size()
vendas_diarias = df.groupby(df["PurchaseDate"].dt.date)["TotalPrice"].sum()
clientes_unicos = df["CustomerID"].nunique()


#dicionario de KPIs
resultados = {
    "KPI": [
        "Receita Total",
        "Média de Vendas por Transação",
        "Total de Transações",
        "Quantidade Total de Produtos Vendidos",
        "Invoice Médio por Cliente",
        "Transações Acima de 5000€",
        "Total de Clientes Únicos"
    ],
    "Valor": [
        receita_total,
        media_vendas,
        total_transacoes,
        total_itens_vendidos,
        invoice_medio_por_cliente ,
        vendas_acima_5000,
        clientes_unicos
    ]
}

#kpis para dataframe
df_resultados = pd.DataFrame(resultados)

#vendas por categoria
vendas_por_categoria_df = vendas_por_categoria.reset_index()
vendas_por_categoria_df.columns = ['Categoria', 'Vendas']

#vendas por categoria e região
vendas_por_categoria_regiao_df = vendas_por_categoria_regiao.reset_index()
vendas_por_categoria_regiao_df.columns = ['Categoria', 'Região', 'Vendas']

#top 5 produtos
top_produtos_df = top_produtos.reset_index()
top_produtos_df.columns = ['ID', 'Regiao', 'Produto','Quantidade Vendida (TOP 5)']

#vendas por método de pagamento
vendas_por_metodo_pagamento_df = vendas_por_metodo_pagamento.reset_index()
vendas_por_metodo_pagamento_df.columns = ['Método de Pagamento', 'Vendas']

#vendas diárias
vendas_diarias_df = vendas_diarias.reset_index()
vendas_diarias_df.columns = ['Data', 'Vendas Diárias']

#guardar em excel
with pd.ExcelWriter(f"KPIs_{data_atual}.xlsx", engine="openpyxl") as writer:
    #kpis gerais (primeira sheet)
    df_resultados.to_excel(writer, sheet_name="KPIS", index=False)

    #sheet vendas diárias
    vendas_diarias_df.to_excel(writer, sheet_name="Vendas Diárias", index=False)
    
    #sheet vendas por categoria
    vendas_por_categoria_df.to_excel(writer, sheet_name="Vendas por Categoria", index=False)
    
    #vendas por categoria e região
    vendas_por_categoria_regiao_df.to_excel(writer, sheet_name="Vendas por Categoria e Região", index=False)
    
    #produtos em uma aba separada
    top_produtos_df.to_excel(writer, sheet_name="Top Produtos", index=False)
    
    #vendas por método de pagamento
    vendas_por_metodo_pagamento_df.to_excel(writer, sheet_name="Vendas por Método de Pagamento", index=False)
    

#let o ficheiro excel para formatar
wb = load_workbook(f"KPIs_{data_atual}.xlsx")
#wb = load_workbook("resultados_vendas.xlsx")

#formatar:
#sheet kpis
ws_metrica = wb["KPIS"]# Ordenar os dados pelo total de quantidade vendida (de forma decrescente) dentro de cada região
ws_metrica.column_dimensions['A'].width = 40
ws_metrica.column_dimensions['B'].width = 40

for i, row in enumerate(df_resultados.values, 2):
    for j, value in enumerate(row):
        cell = ws_metrica.cell(row=i, column=j+1, value=value)
        if i == 1:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

#sheet vendas por categoria
ws_vendas = wb["Vendas por Categoria"]
ws_vendas.column_dimensions['A'].width = 20
ws_vendas.column_dimensions['B'].width = 20

for i, row in enumerate(vendas_por_categoria_df.values, 2):
    for j, value in enumerate(row):
        cell = ws_vendas.cell(row=i, column=j+1, value=value)
        if i == 1:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

#sheet vendas por categoria e regiao
ws_vendas_categoria_regiao = wb["Vendas por Categoria e Região"]
ws_vendas_categoria_regiao.column_dimensions['A'].width = 20
ws_vendas_categoria_regiao.column_dimensions['B'].width = 20
ws_vendas_categoria_regiao.column_dimensions['C'].width = 20


for i, row in enumerate(vendas_por_categoria_regiao_df.values, 2):
    for j, value in enumerate(row):
        cell = ws_vendas_categoria_regiao.cell(row=i, column=j+1, value=value)
        if i == 1:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

#sheet top produtos
ws_top_produtos = wb["Top Produtos"]
ws_top_produtos.column_dimensions['A'].width = 10
ws_top_produtos.column_dimensions['B'].width = 20
ws_top_produtos.column_dimensions['C'].width = 20
ws_top_produtos.column_dimensions['D'].width = 30


for i, row in enumerate(top_produtos_df.values, 2):
    for j, value in enumerate(row):
        cell = ws_top_produtos.cell(row=i, column=j+1, value=value)
        if i == 1:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

#sheet vendas por metodo de pagamento
ws_vendas_metodo_pagamento = wb["Vendas por Método de Pagamento"]
ws_vendas_metodo_pagamento.column_dimensions['A'].width = 30
ws_vendas_metodo_pagamento.column_dimensions['B'].width = 20

for i, row in enumerate(vendas_por_metodo_pagamento_df.values, 2):
    for j, value in enumerate(row):
        cell = ws_vendas_metodo_pagamento.cell(row=i, column=j+1, value=value)
        if i == 1:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

#sheet vendas por dia
ws_vendas_diarias = wb["Vendas Diárias"]
ws_vendas_diarias.column_dimensions['A'].width = 20
ws_vendas_diarias.column_dimensions['B'].width = 20


for i, row in enumerate(vendas_diarias_df.values, 2):
    for j, value in enumerate(row):
        cell = ws_vendas_diarias.cell(row=i, column=j+1, value=value)
        if i == 1:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

#save
wb.save(f"KPIs_{data_atual}.xlsx")
print("Novo arquivo com formatação foi salvo com sucesso!")
